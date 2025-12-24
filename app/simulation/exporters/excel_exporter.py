"""
Excel exporter for agent personas.

Creates multi-sheet Excel files with formatted data:
- Summary: Key metrics overview
- All Attributes: Complete 20+ attribute table
- Backstories: Full narrative descriptions
- Behavioral Profile: Psychological trait scores
- Shopping Patterns: Sample behavior descriptions
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils.dataframe import dataframe_to_rows

from app.simulation.models.persona import AgentPersona

logger = logging.getLogger(__name__)


class PersonaExcelExporter:
    """
    Export agent personas to formatted Excel file.

    Usage:
        exporter = PersonaExcelExporter(personas)
        exporter.export("data/personas/personas_20241224.xlsx")
    """

    # Style definitions
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    BORDER_THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, personas: List[AgentPersona]):
        self.personas = personas

    def export(self, filepath: Path) -> None:
        """
        Create multi-sheet Excel file with all persona data.

        Args:
            filepath: Path where Excel file will be saved
        """
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create all sheets
        self._create_summary_sheet(wb)
        self._create_attributes_sheet(wb)
        self._create_backstories_sheet(wb)
        self._create_behavioral_sheet(wb)
        self._create_patterns_sheet(wb)

        # Save
        wb.save(filepath)
        logger.info(f"Excel exported to: {filepath}")

    def append(self, filepath: Path) -> None:
        """
        Append personas to an existing Excel file.

        Args:
            filepath: Path to existing Excel file to append to
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {filepath}")

        wb = load_workbook(filepath)

        # Get the max agent_id from existing data to assign new IDs
        existing_max_id = self._get_max_agent_id(wb)
        start_id_num = existing_max_id + 1

        logger.info(f"Appending {len(self.personas)} personas starting from agent_{start_id_num:03d}")

        # Reassign agent IDs for new personas
        for i, persona in enumerate(self.personas):
            persona.agent_id = f"agent_{start_id_num + i:03d}"
            persona.generated_at = datetime.utcnow().isoformat()

        # Append to each sheet
        self._append_to_summary_sheet(wb)
        self._append_to_attributes_sheet(wb)
        self._append_to_backstories_sheet(wb)
        self._append_to_behavioral_sheet(wb)
        self._append_to_patterns_sheet(wb)

        # Save
        wb.save(filepath)
        logger.info(f"Appended {len(self.personas)} personas to: {filepath}")

    def _get_max_agent_id(self, wb: Workbook) -> int:
        """
        Extract the highest agent_id number from existing file.

        Args:
            wb: The workbook to read from

        Returns:
            The highest agent_id number found (0 if none)
        """
        ws = wb["Summary"]
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], str) and row[0].startswith("agent_"):
                try:
                    num = int(row[0].split("_")[1])
                    max_id = max(max_id, num)
                except (ValueError, IndexError):
                    continue
        return max_id

    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet with key metrics overview."""
        ws = wb.create_sheet("Summary")

        # Headers
        headers = [
            "Agent ID",
            "Age",
            "Gender",
            "Income",
            "Location",
            "Price Sensitivity",
            "Brand Loyalty",
            "Coupon Affinity",
            "Primary Categories",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Data rows
        for row_idx, persona in enumerate(self.personas, start=2):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.gender).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=4, value=persona.demographics.income_bracket).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=5, value=persona.demographics.location_region).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=6, value=persona.behavioral_traits.price_sensitivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=7, value=persona.behavioral_traits.brand_loyalty).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=8, value=persona.coupon_behavior.coupon_affinity).border = self.BORDER_THIN
            ws.cell(
                row=row_idx,
                column=9,
                value=", ".join(persona.shopping_preferences.preferred_categories[:3]),
            ).border = self.BORDER_THIN

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _create_attributes_sheet(self, wb: Workbook) -> None:
        """Create detailed attributes sheet with 20+ columns."""
        ws = wb.create_sheet("All Attributes")

        # All attribute columns
        headers = [
            "agent_id",
            # Demographics
            "age",
            "age_group",
            "gender",
            "income_bracket",
            "household_size",
            "has_children",
            "location_region",
            # Behavioral
            "price_sensitivity",
            "brand_loyalty",
            "impulsivity",
            "tech_savviness",
            # Shopping
            "preferred_categories",
            "weekly_budget",
            "shopping_frequency",
            "avg_cart_value",
            # Temporal - days
            "pref_day_weekday",
            "pref_day_saturday",
            "pref_day_sunday",
            # Temporal - times
            "pref_time_morning",
            "pref_time_afternoon",
            "pref_time_evening",
            # Coupon
            "coupon_affinity",
            "deal_seeking_behavior",
            # Metadata
            "generation_model",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Data rows
        for row_idx, persona in enumerate(self.personas, start=2):
            data = persona.to_dict()
            for col, header in enumerate(headers, 1):
                value = data.get(header, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = self.BORDER_THIN

                # Align numeric values right
                if header in ["price_sensitivity", "brand_loyalty", "impulsivity", "tech_savviness",
                              "weekly_budget", "avg_cart_value", "coupon_affinity",
                              "pref_day_weekday", "pref_day_saturday", "pref_day_sunday",
                              "pref_time_morning", "pref_time_afternoon", "pref_time_evening"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

    def _create_backstories_sheet(self, wb: Workbook) -> None:
        """Create backstories sheet with full narrative descriptions."""
        ws = wb.create_sheet("Backstories")

        # Headers
        headers = ["Agent ID", "Age", "Gender", "Backstory"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Data rows
        for row_idx, persona in enumerate(self.personas, start=2):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.gender).border = self.BORDER_THIN
            cell = ws.cell(row=row_idx, column=4, value=persona.backstory)
            cell.border = self.BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 80  # Wide column for backstory

    def _create_behavioral_sheet(self, wb: Workbook) -> None:
        """Create behavioral profile sheet with psychological scores."""
        ws = wb.create_sheet("Behavioral Profile")

        # Headers
        headers = [
            "Agent ID",
            "Age",
            "Price Sensitivity",
            "Brand Loyalty",
            "Impulsivity",
            "Tech Savviness",
            "Coupon Affinity",
            "Deal Seeking",
            "Shopping Frequency",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Data rows
        for row_idx, persona in enumerate(self.personas, start=2):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.behavioral_traits.price_sensitivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=4, value=persona.behavioral_traits.brand_loyalty).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=5, value=persona.behavioral_traits.impulsivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=6, value=persona.behavioral_traits.tech_savviness).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=7, value=persona.coupon_behavior.coupon_affinity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=8, value=persona.coupon_behavior.deal_seeking_behavior).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=9, value=persona.shopping_preferences.shopping_frequency).border = self.BORDER_THIN

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column].width = adjusted_width

    def _create_patterns_sheet(self, wb: Workbook) -> None:
        """Create shopping patterns sheet with sample behaviors."""
        ws = wb.create_sheet("Shopping Patterns")

        # Headers
        headers = ["Agent ID", "Age", "Income", "Shopping Patterns"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_THIN

        # Data rows
        for row_idx, persona in enumerate(self.personas, start=2):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.income_bracket).border = self.BORDER_THIN

            # Join patterns with newlines
            patterns_text = "\n".join(f"• {p}" for p in persona.sample_shopping_patterns)
            cell = ws.cell(row=row_idx, column=4, value=patterns_text)
            cell.border = self.BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 80

    # ========== Append Methods ==========

    def _append_to_summary_sheet(self, wb: Workbook) -> None:
        """Append personas to the Summary sheet."""
        ws = wb["Summary"]
        start_row = ws.max_row + 1

        for row_idx, persona in enumerate(self.personas, start=start_row):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.gender).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=4, value=persona.demographics.income_bracket).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=5, value=persona.demographics.location_region).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=6, value=persona.behavioral_traits.price_sensitivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=7, value=persona.behavioral_traits.brand_loyalty).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=8, value=persona.coupon_behavior.coupon_affinity).border = self.BORDER_THIN
            categories = ", ".join(persona.shopping_preferences.preferred_categories[:3])
            ws.cell(row=row_idx, column=9, value=categories).border = self.BORDER_THIN

    def _append_to_attributes_sheet(self, wb: Workbook) -> None:
        """Append personas to the All Attributes sheet."""
        ws = wb["All Attributes"]
        start_row = ws.max_row + 1

        # Get column headers from row 1
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(header)

        # Map headers to persona data
        header_to_data = {
            "agent_id": lambda p: p.agent_id,
            "age": lambda p: p.demographics.age,
            "age_group": lambda p: p.demographics.age_group,
            "gender": lambda p: p.demographics.gender,
            "income_bracket": lambda p: p.demographics.income_bracket,
            "household_size": lambda p: p.demographics.household_size,
            "has_children": lambda p: p.demographics.has_children,
            "location_region": lambda p: p.demographics.location_region,
            "price_sensitivity": lambda p: p.behavioral_traits.price_sensitivity,
            "brand_loyalty": lambda p: p.behavioral_traits.brand_loyalty,
            "impulsivity": lambda p: p.behavioral_traits.impulsivity,
            "tech_savviness": lambda p: p.behavioral_traits.tech_savviness,
            "preferred_categories": lambda p: ", ".join(p.shopping_preferences.preferred_categories),
            "weekly_budget": lambda p: p.shopping_preferences.weekly_budget,
            "shopping_frequency": lambda p: p.shopping_preferences.shopping_frequency,
            "avg_cart_value": lambda p: p.shopping_preferences.avg_cart_value,
            "pref_day_weekday": lambda p: p.temporal_patterns.preferred_days.get("weekday", 0.0),
            "pref_day_saturday": lambda p: p.temporal_patterns.preferred_days.get("saturday", 0.0),
            "pref_day_sunday": lambda p: p.temporal_patterns.preferred_days.get("sunday", 0.0),
            "pref_time_morning": lambda p: p.temporal_patterns.preferred_times.get("morning", 0.0),
            "pref_time_afternoon": lambda p: p.temporal_patterns.preferred_times.get("afternoon", 0.0),
            "pref_time_evening": lambda p: p.temporal_patterns.preferred_times.get("evening", 0.0),
            "coupon_affinity": lambda p: p.coupon_behavior.coupon_affinity,
            "deal_seeking_behavior": lambda p: p.coupon_behavior.deal_seeking_behavior,
            "generation_model": lambda p: p.generation_model or "",
        }

        for row_idx, persona in enumerate(self.personas, start=start_row):
            for col, header in enumerate(headers, 1):
                if header in header_to_data:
                    value = header_to_data[header](persona)
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = self.BORDER_THIN

                    # Align numeric values right
                    if header in ["price_sensitivity", "brand_loyalty", "impulsivity", "tech_savviness",
                                  "weekly_budget", "avg_cart_value", "coupon_affinity",
                                  "pref_day_weekday", "pref_day_saturday", "pref_day_sunday",
                                  "pref_time_morning", "pref_time_afternoon", "pref_time_evening"]:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

    def _append_to_backstories_sheet(self, wb: Workbook) -> None:
        """Append personas to the Backstories sheet."""
        ws = wb["Backstories"]
        start_row = ws.max_row + 1

        for row_idx, persona in enumerate(self.personas, start=start_row):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.gender).border = self.BORDER_THIN
            cell = ws.cell(row=row_idx, column=4, value=persona.backstory)
            cell.border = self.BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    def _append_to_behavioral_sheet(self, wb: Workbook) -> None:
        """Append personas to the Behavioral Profile sheet."""
        ws = wb["Behavioral Profile"]
        start_row = ws.max_row + 1

        for row_idx, persona in enumerate(self.personas, start=start_row):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.behavioral_traits.price_sensitivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=4, value=persona.behavioral_traits.brand_loyalty).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=5, value=persona.behavioral_traits.impulsivity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=6, value=persona.behavioral_traits.tech_savviness).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=7, value=persona.coupon_behavior.coupon_affinity).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=8, value=persona.coupon_behavior.deal_seeking_behavior).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=9, value=persona.shopping_preferences.shopping_frequency).border = self.BORDER_THIN

    def _append_to_patterns_sheet(self, wb: Workbook) -> None:
        """Append personas to the Shopping Patterns sheet."""
        ws = wb["Shopping Patterns"]
        start_row = ws.max_row + 1

        for row_idx, persona in enumerate(self.personas, start=start_row):
            ws.cell(row=row_idx, column=1, value=persona.agent_id).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=2, value=persona.demographics.age).border = self.BORDER_THIN
            ws.cell(row=row_idx, column=3, value=persona.demographics.income_bracket).border = self.BORDER_THIN

            patterns_text = "\n".join(f"• {p}" for p in persona.sample_shopping_patterns)
            cell = ws.cell(row=row_idx, column=4, value=patterns_text)
            cell.border = self.BORDER_THIN
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
